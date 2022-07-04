import xml.etree.ElementTree as ET
from openpyxl import load_workbook



date = [] #date
dtpv = [] #DTPV
time = []
# first level


coordinate_lenght = [] #COORD_L 
coordinate_weight = [] #COORD_W 
house = [] #house
object_dtp = [] #k_ul
osv = [] #osv
street = [] #street
road_status = [] #s_pch
weather = []#spog
#second level


injury = [] #s_T
#third level


tree = ET.parse('Список карточек ДТП.xml')
root = tree.getroot()
for items in root:
    for tab in items:
        if 'date' in tab.tag:
            date.append(tab.text)
        if 'DTPV' in tab.tag:
            dtpv.append(tab.text)
        if 'time' in tab.tag:
            time.append(tab.text)
        for info_dtp in tab:
            if 'COORD_L' in info_dtp.tag:
                coordinate_lenght.append(info_dtp.text)
            if 'COORD_W' in info_dtp.tag:
                coordinate_weight.append(info_dtp.text)
            if 'k_ul' in info_dtp.tag:
                object_dtp.append(info_dtp.text)
            if 'osv' in info_dtp.tag:
                osv.append(info_dtp.text)
            if 'house' in info_dtp.tag:
                house.append(info_dtp.text)
            if 'street' in info_dtp.tag:
                street.append(info_dtp.text)
            if 's_pch' in info_dtp.tag:
                road_status.append(info_dtp.text)
            if 'spog' in info_dtp.tag:
                weather.append(info_dtp.text)
            for ts_info in info_dtp:
                if 's_T' in ts_info.tag:
                    injury.append(ts_info.text)
                



wb = load_workbook(filename='Лист Microsoft Excel.xlsx')
ws = wb['Лист1']
column = 1
for rows in range(len(date)):
    ws.cell(row=rows+1, column=column, value=date[rows])
for rows in range(len(time)):
    ws.cell(row=rows+1, column=column+1, value=time[rows])
for rows in range(len(coordinate_lenght)):
    ws.cell(row=rows+1, column=column+11, value=coordinate_lenght[rows])
for rows in range(len(coordinate_weight)):
    ws.cell(row=rows+1, column=column+10, value=coordinate_weight[rows])
for rows in range(len(object_dtp)):
    ws.cell(row=rows+1, column=column+7, value=object_dtp[rows])
for rows in range(len(osv)):
    ws.cell(row=rows+1, column=column+2, value=osv[rows])
for rows in range(len(weather)):
    ws.cell(row=rows+1, column=column+3, value=weather[rows])
for rows in range(len(injury)):
    ws.cell(row=rows+1, column=column+9, value=injury[rows])
for rows in range(len(dtpv)):
    ws.cell(row=rows+1, column=column+8, value=dtpv[rows])
for rows in range(len(house)):
    ws.cell(row=rows+1, column=column+6, value=house[rows])
for rows in range(len(street)):
    ws.cell(row=rows+1, column=column+5, value=street[rows])
for rows in range(len(road_status)):
    ws.cell(row=rows+1, column=column+4, value=road_status[rows])
wb.save('Лист Microsoft Excel.xlsx')
